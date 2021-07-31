#!/usr/bin/env python
# coding: utf-8
import tkinter
#import openpyxl
from openpyxl import load_workbook
from tkinter import *
#import datetime
#from datetime import datetime
#=================
#==============




wb = load_workbook('A:\Python\Report.xlsx')
sheet = wb.active
max_col = sheet.max_column
#max_ro = sheet.max_row
#now = datetime.now()
#date_time = now.strftime("%Y-%d-%m %H:%M:%S")
#date_time1 = datetime.strpttime(date_time, "%Y-%d-%m %H:%M:%S")
#now_time = now.strftime("%H:%M:%S")

inprocount = 0
completedcount = 0
scheduledcount = 0
rescheduledcount = 0
holdcount = 0
failedcount = 0
cancelledcount = 0
notyetapprovedcount = 0
otherscount = 0
firstequation = ""
secondequation =""
pchange=0
npchange=0
fshift=0
s_change=""


def inpro():
    
    global inprocount
    global completedcount
    global scheduledcount
    global rescheduledcount
    global holdcount
    global failedcount
    global cancelledcount
    global notyetapprovedcount
    global otherscount

    inprocount = 0
    completedcount = 0
    scheduledcount = 0
    rescheduledcount = 0
    holdcount = 0
    failedcount = 0
    cancelledcount = 0
    notyetapprovedcount = 0
    otherscount = 0
        
    for i in range(2, (sheet.max_row)+1):
        cell_obj = sheet.cell(row=i, column=9)
        if(cell_obj.value == "In progress with issues" or cell_obj.value == "In progress without issues"):
            inprocount+=1
        elif(cell_obj.value == "Completed with issues" or cell_obj.value == "Completed without issues"):
            completedcount+=1
        elif(cell_obj.value == "Scheduled"):
            scheduledcount+=1
        elif(cell_obj.value == "Rescheduled"):
            rescheduledcount+=1
        elif(cell_obj.value == "ON-HOLD"):
            holdcount+=1
        elif(cell_obj.value == "Failed"):
            failedcount+=1
        elif(cell_obj.value == "Cancelled"):
            cancelledcount+=1
        elif(cell_obj.value == "Not Yet Approved"):
            notyetapprovedcount+=1
        else:
            otherscount+=1
            
    inproequation.set(inprocount)
    completedequation.set(completedcount)
    scheduledequation.set(scheduledcount)
    rescheduledequation.set(rescheduledcount)
    holdequation.set(holdcount)
    failedequation.set(failedcount)
    cancelledequation.set(cancelledcount)
    notyetapprovedequation.set(notyetapprovedcount)
    othersequation.set(otherscount)

    


def clear():
    inproequation.set("")
    completedequation.set("")
    scheduledequation.set("")
    rescheduledequation.set("")
    holdequation.set("")
    failedequation.set("")
    cancelledequation.set("")
    notyetapprovedequation.set("")
    othersequation.set("")
    prodequetion.set("")
    nonprodequetion.set("")
    searchequation.set("")
    inprocount = 0
    completedcount = 0
    scheduledcount = 0
    rescheduledcount = 0
    holdcount = 0
    failedcount = 0
    cancelledcount = 0
    notyetapprovedcount = 0
    otherscount = 0
    npchange =0
    pchange=0

def display():
    #global firstequation
    data1 = sheet['A1':'K200']
    print("\nAll Changes\n")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i1.value == None:
            break
        else:
            print("{0} ------ {2} -----------{3} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def compchanges():
    data1 = sheet['A1':'K200']
    print("\nChanges Completed\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "Completed with issues" or i9.value == "Completed without issues":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def inprochanges():
    print("Changes Inprogress\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    data1 = sheet['A1':'K200']
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "In progress with issues" or i9.value == "In progress without issues":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def scheduledchanges():
    data1 = sheet['A1':'K200']
    print("\nScheduled Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "Scheduled":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def rescheduledchanges():
    data1 = sheet['A1':'K200']
    print("\nRescheduled Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "Rescheduled":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))            

def onholdchanges():
    data1 = sheet['A1':'K200']
    print("\nOn-Hold Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "ON-HOLD":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def failedchanges():
    data1 = sheet['A1':'K200']
    print("\nFailed Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "Failed":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def cancelledchanges():
    data1 = sheet['A1':'K200']
    print("\nCancelled Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "Cancelled":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def notapprovedchanges():
    data1 = sheet['A1':'K200']
    print("\nNot Approved Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == "Not Yet Approved":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def otherchanges():
    data1 = sheet['A1':'K200']
    print("\nNot Assigned Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i9.value == None and i1.value != None:
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
"""    
    rows = []
    for i in range(1, max_col):
        cols = []
        for j in range(4):
            e = Entry(relief=GROOVE)

            e.grid(row=i, column=j, sticky=NSEW)

            e.insert(END, '%d.%d' % sheet.cell(row=i, column=j))

            cols.append(e)

        rows.append(cols)

   
            

     cell_obj = sheet.cell(row=j, column=i)
        print(cell_obj.value, end="  ")
        from tkinter import *



for i in range(5):

    

    

    data2 = sheet['A3':'K3']
    for j1, j2, j3, j4, j5, j6, j7, j8, j9, j10, j11 in data2:
        secondequation.set("{0:1} - {10:1} - {2:1} - {4:1} - {8:1} - {6:1}".format(j1.value, j2.value, j3.value, j4.value, j5.value,\
        j6.value, j7.value, j8.value, j9.value, j10.value, j11.value))        
"""     
def prodchanges():
    global pchange
    pchange = 0
    data1 = sheet['A1':'K200']
    print("\nProd Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i3.value == 'Prod' or i3.value == 'prod':
            pchange+=1
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
    prodequetion.set(pchange)

    
def nonprodchanges():
    global npchange
    npchange =0
    data1 = sheet['A1':'K200']
    print("\nNonProd Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i3.value != 'Prod' and i1.value != None:
            npchange+=1
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
    nonprodequetion.set(npchange)
    

def coming_changes():
    global fshift
    data1 = sheet['A1':'K200']
    print("\nFirst Shift Changes\n")
    
    for i in range(2, ):
        now1 = str(sheet.cell(row=i, column=4).value)
        dt_object1 = datetime.strptime(now1, "%Y-%m-%d %H:%M:%S")
        time_obj = datetime.strftime(dt_object1, "%H:%M:%S")
        if  dt_object1 < date_time:
            print("Old change")
        else:
            print("New change")
    
        
        
    
         

'''
    
         
        for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        mystring = i4.value
        mystring1 = mystring.strptime("%Y-%m-%d %H:%M:%S")
        #mystring2 = datetime.strftime(mystring1, "%Y-%m-%d %H:%M:%S")
        #mystring2 = mystring1.strftime("%Y-%d-%m")

        print(mystring1)
        
        date_obj = datetime.strftime(mystring, "%Y-%m-%d %H:%M%S")
        my_time = datetime.strftime(date_obj, "%H:%M%S")
        strttime = datetime.strftime(mystring, "%H:%M%S")
        if strttime >= '6:30:00' and strttime <= '14:30:00':
            fshift+=1
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
'''                                                                                              #i5.value, i6.value, i7.value, i8.value,\
                                                                                                          #i9.value, i10.value, i11.value))
    #nonprodequetion.set(npchange)


def searching_change():
    global s_change
    global searchequation
    data1 = sheet['A1':'K200']
    print("\nSearched Result")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    s_change = searchequation.get()
    #print(s_change)
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i1.value == str(s_change):
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
            break
    else:
        print("No result found for the searched item")
'''
def bar_chart():
    chart = BarChart()  
    values = Reference(worksheet=sheet,  
                 min_row=2,  
                 max_row=73,  
                 min_col=9,  
                 max_col=9)  
  
    chart.add_data(values, titles_from_data=True)  
    sheet.add_chart(chart, "E2")  
'''


def wmk_display():
    data1 = sheet['A1':'K200']
    print("\nWMK Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Wellmark (WMK)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def bcn_display():
    data1 = sheet['A1':'K200']
    print("\nBCN Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Blue Care Network (BCN)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def emb_display():
    data1 = sheet['A1':'K200']
    print("\nEMB Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Emblem Health (EMB)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def avm_display():
    data1 = sheet['A1':'K200']
    print("\nAVM/MMM Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "InnovaCare (AVM)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))                                                                                                          
                                                                                                
def vph_display():
    data1 = sheet['A1':'K200']
    print("\nVPH Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Virginia Premier Health (VPH)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def wha_display():
    data1 = sheet['A1':'K200']
    print("\nWHA Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Western Health Advantage (WHA)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def hsn_display():
    data1 = sheet['A1':'K200']
    print("\nHSN Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Health Services for Children with Special Needs (HSN)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def dhp_display():
    data1 = sheet['A1':'K200']
    print("\nDHP Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Dean Health Plan (DHP)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def lac_display():
    data1 = sheet['A1':'K200']
    print("\nLAC Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "LA Care Health Plan (LAC)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def swh_display():
    data1 = sheet['A1':'K200']
    print("\nSWH Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Senior Whole Health (SWH)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def nhp_display():
    data1 = sheet['A1':'K200']
    print("\nAWHP/NHP Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Allways Health Partners (AWHP/NHP)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def ich_display():
    data1 = sheet['A1':'K200']
    print("\nICH Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Independent Care Health Plan (ICH)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))
def vhp_display():
    data1 = sheet['A1':'K200']
    print("\nVHP Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Valley Health Plan (VHP)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def coa_display():
    data1 = sheet['A1':'K200']
    print("\nCOA Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "Colorado Access Health (COA)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

def bri_display():
    data1 = sheet['A1':'K200']
    print("\nBRI Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value == "BCBS of Rhode Island (BRI)":
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))


def otherclient_display():
    data1 = sheet['A1':'K200']
    print("\nOther Client Changes\n")
    print("Ticket ID ------ Environment ----- Planned End Date ----------- Status ------------ Change Coordinator --------- Company")
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
        if i11.value != "Wellmark (WMK)" and i11.value != "Blue Care Network (BCN)" and i11.value != "Emblem Health (EMB)" and i11.value != "InnovaCare (AVM)"\
           and i11.value != "Virginia Premier Health (VPH)" and i11.value != "Western Health Advantage (WHA)"\
           and i11.value != "Health Services for Children with Special Needs (HSN)" and i11.value != "Dean Health Plan (DHP)" \
           and i11.value != "LA Care Health Plan (LAC)" and i11.value != "Senior Whole Health (SWH)" and i11.value != "Allways Health Partners (AWHP/NHP)"\
           and i11.value != "Independent Care Health Plan (ICH)" and i11.value != "Valley Health Plan (VHP)" and i11.value != "Colorado Access Health (COA)"\
           and i11.value != "BCBS of Rhode Island (BRI)" and i1.value != None:
                     
            print("{0} ------ {2} -----------{4} ------------{8} --------------{6} ----------{10}".format(i1.value,\
                                                                                                          i2.value, i3.value, i4.value,\
                                                                                                          i5.value, i6.value, i7.value, i8.value,\
                                                                                                          i9.value, i10.value, i11.value))

if __name__ == "__main__":
    root = Tk()
    root.configure(background='slateblue')
    root.geometry("1000x600")
    root.title("Changes Summery")
    

    inproequation = StringVar()
    completedequation = StringVar()
    scheduledequation = StringVar()
    rescheduledequation = StringVar()
    holdequation = StringVar()
    failedequation = StringVar()
    cancelledequation = StringVar()
    notyetapprovedequation = StringVar()
    othersequation = StringVar()
    firstequation = StringVar()
    secondequation = StringVar()
    prodequetion = StringVar()
    nonprodequetion = StringVar()
    searchequation = StringVar()

    
    count = Label(root, text="Total Count", bg="white", fg='blue')
    inprogress = Button(root, text="In Progress", bg="black", command = inprochanges, fg='yellow', relief=GROOVE)
    completed = Button(root, text="Completed", bg="black", command = compchanges,fg='yellow', relief=GROOVE)
    scheduled = Button(root, text="Scheduled", bg="black", command = scheduledchanges, fg='yellow', relief=GROOVE)
    rescheduled = Button(root, text="Rescheduled", bg="black", command = rescheduledchanges, fg='yellow', relief=GROOVE)
    hold = Button(root, text="Hold", bg="black", command = onholdchanges, fg='yellow', relief=GROOVE)
    failed = Button(root, text="Failed", bg="black", command = failedchanges,fg='yellow', relief=GROOVE)
    cancelled = Button(root, text="Cancelled", bg="black", command = cancelledchanges, fg='yellow', relief=GROOVE)
    notyetapproved = Button(root, text="Not approved", bg="black", command = notapprovedchanges, fg='yellow', relief=GROOVE)
    others = Button(root, text="Others", bg="black", command = otherchanges, fg='yellow', relief=GROOVE)

    countmessage = Message(root, text=(sheet.max_row)-1, bd=2, relief=RIDGE, padx=2, font=('times', 18, 'italic'))
    inprogressmessage = Entry(root, textvariable=inproequation)
    completedmessage= Entry(root, textvariable= completedequation)
    scheduledmessage = Entry(root, textvariable=scheduledequation)
    rescheduledmessage = Entry(root, textvariable=rescheduledequation)
    holdmessage = Entry(root, textvariable= holdequation)
    failedmessage = Entry(root, textvariable= failedequation)
    cancelledmessage = Entry(root, textvariable=cancelledequation)
    notyetapprovedmessage = Entry(root, textvariable=notyetapprovedequation)
    othersmessage = Entry(root, textvariable = othersequation)
    
    count.place(x=10, y=30)
    inprogress.place(x=100, y=30)
    completed.place(x=190, y=30)
    scheduled.place(x=280, y=30)
    rescheduled.place(x=370, y=30)
    hold.place(x=460, y=30)
    failed.place(x=550, y=30)
    cancelled.place(x=640, y=30)
    notyetapproved.place(x=730, y=30)
    others.place(x=820, y=30)    
            
    countmessage.place(x=10, y=60)
    inprogressmessage.place(x=100, y=60 )
    completedmessage.place(x=190, y=60)
    scheduledmessage.place(x=280, y=60)
    rescheduledmessage.place(x=370, y=60)
    holdmessage.place(x=460, y=60)
    failedmessage.place (x=550, y=60)
    cancelledmessage.place(x=640, y=60)
    notyetapprovedmessage.place(x=730, y=60)
    othersmessage.place(x=820, y=60) 

    submit = Button(root, text="Display Data", command=inpro, fg='lavender', bg='midnightblue', height=2, width=10, relief=RAISED)
    submit.place(x=30, y=150)

    reset = Button(root, text="Reset All", command=clear, fg='lavender', bg='darkgreen', height=2, width=10, relief=RAISED)
    reset.place(x=30, y=200)

    dispbutton = Button(root, text="Display All Changes", command=display, width=30)
    prodbutton = Button(root, text="Prod", command=prodchanges)
    nonprodbutton = Button(root, text="NonProd", command=nonprodchanges)
    comingchangesbutton = Button(root, text="Coming Changes", command=coming_changes)
    previouschangesbutton =Button(root,text="Previous Changes")
    search_button = Button(root, text="Search Change", command=searching_change)
    #barchart_button = Button(root, text="Bar chart", command=bar_chart)
    excelMessage = Message(root, width=1000, text="Column details in Sheet should be in below order \nTicket ID, Description, Environment Affected,\
 Planned Start Date, Planned End Date, Opened by, Change Coordinator, Action Required, Status, Ritm State, Company")

    wmkbutton = Button(root, text="WMK", command=wmk_display)
    embbutton = Button(root, text="EMB", command=emb_display)
    bcnbutton = Button(root, text="BCN", command=bcn_display)
    avmbutton = Button(root, text="AVM", command=avm_display)
    vphbutton = Button(root, text="VPH", command=vph_display)
    whabutton = Button(root, text="WHA", command=wha_display)
    hsnbutton = Button(root, text="HSN", command=hsn_display)
    dhpbutton = Button(root, text="DHP", command=dhp_display)
    lacbutton = Button(root, text="LAC", command=lac_display)
    swhbutton = Button(root, text="SWH", command=swh_display)
    nhpbutton = Button(root, text="NHP", command=nhp_display)
    ichbutton = Button(root, text="ICH", command=ich_display)
    vhpbutton = Button(root, text="VHP", command=vhp_display)
    coabutton = Button(root, text="COA", command=coa_display)
    bributton = Button(root, text="BRI", command=bri_display)
    remainingclientbutton = Button(root, text="Other_Client", command=otherclient_display)

    prodentry = Entry(root, textvariable=prodequetion)
    nonprodentry = Entry(root, textvariable=nonprodequetion)
    searchchange_entry = Entry(root, textvariable =searchequation)
    

    """
    dispmessage = Message(root, textvariable=firstequation, width=800)
    descentry = Entry(root, textvariable=secondequation, width=100)
    
    enventry = Entry(root)
    starttentry = Entry(root)
    statusentry = Entry(root)
    """

    dispbutton.place(x=30, y=330)
    prodbutton.place(x=30, y=370)
    nonprodbutton.place(x=30, y=420)
    #comingchangesbutton.place(x=30, y=500)
    #previouschangesbutton.place(x=30, y=550)
    search_button.place(x=350, y=200)
    #barchart_button.place(x=400, y=300)
    

    prodentry.place(x=100, y=370)
    nonprodentry.place(x=100, y=420)
    searchchange_entry.place(x=350, y=230)
    excelMessage.place(x=10, y=520)
    wmkbutton.place(x=700, y=150)
    embbutton.place(x=700, y=200)
    bcnbutton.place(x=700, y=250)
    avmbutton.place(x=700, y=300)
    vphbutton.place(x=700, y=350)
    hsnbutton.place(x=750, y=150)
    dhpbutton.place(x=750, y=200)
    lacbutton.place(x=750, y=250)
    swhbutton.place(x=750, y=300)
    nhpbutton.place(x=750, y=350)
    whabutton.place(x=800, y=150)
    ichbutton.place(x=800, y=200)
    vhpbutton.place(x=800, y=250)
    coabutton.place(x=800, y=300)
    bributton.place(x=800, y=350)
    remainingclientbutton.place(x=730, y=400)
    """
    dispmessage.place(x=130, y=260)
    #descentry.place(x=130, y=360)
    
    enventry.place(x=330, y=260)
    starttentry.place(x=430, y=260)
    statusentry.place(x=530, y=260)
    """
    root.mainloop()
