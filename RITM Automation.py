from tkinter import *
import tkinter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import datetime
from datetime import datetime



#======Loading Excelsheet=======
wb = load_workbook('Daily RITM Reporting.xlsx')
sheet = wb.active

def saveexcel():
    wb.save('Daily RITM Reporting.xlsx')

#=====Date and time======
#today = datetime.date.today()
now = datetime.now()
current_time = now.strftime("%H-%M-%S")
dt_string = now.strftime("%Y-%m-%d %H:%M:%S")


ClientName = ["Ascension Care Management LLC (ABS)","Advanced Benefit Solutions LLC - QL (ABS)","Allways Health Partners (AWHP/NHP)",\
              "BCBS of Rhode Island (BRI)","BCBS North Carolina (BNC)", "Blue Care Network (BCN)","Bright Health Plan (BRT)",\
              "Centers Plan for Healthy Living (CPL)","CIS Healthcare Operations Center (CISHOC)","Colorado Access Health (COA)",\
              "Community First Health Plan (CFH)","County of Ventura (COV)","Dean Health Plan (DHP)","Denver Health Medical Plan (DHH)",\
              "Emblem Health (EMB)","Fallon Community Health Plan (FAL)","Gateway health plan (GHP)","Hamaspik Choice Inc (HCI)",\
              "Healthcare Highways, Inc. - QL (HCH)","Health Services for Children with Special Needs (HSN)","Hawaii Medical Service Assoc (HMS)",\
              "Independent Care Health Plan (ICH)","InnovaCare (AVM)","Keystone Mercy Health Plan (KMH)","Memorial Hermann Health (MHH)",
              "LA Care Health Plan (LAC)","PacificSource Health Plans (PFS)","Physicians Health Plan Shared Services (PHP)","Prominence Health Plan (PRH)",\
              "QualChoice of Arkansas (QCA)","Senior Whole Health (SWH)","Texas Children Health Plan (TXC)","United HealthCare ClaimSphere (UHG)",\
              "UnitedHealth Group Complex Medical Conditions (CMC)","University Health Alliance (UHA)","Valley Health Plan (VHP)","Virginia Premier Health (VPH)",\
              "Vibra Health Plan (VIB)","Viva Health Inc. (VHI)","Wellmark (WMK)"]

TotalRitm = 0
OpenCount = 0
PendingCount=0
WipCount=0
OthersStateCount=0
CriticalCount=0
HighCount=0
MediumCount=0
LowCount=0
SearchRitmRef =""
WorknoteRitmRef =""
WorknoteRitmRef1=""
z=0

#=====Clear, Ttal Count, Exit functions ======
def clear():
    OpenRitmEqu.set("")
    PendingRitmEqu.set("")
    WipRitmEqu.set("")
    OtherStateRitmEqu.set("")
    CriticalRitmEqu.set("")
    HighRitmEqu.set("")
    MediumRitmEqu.set("")
    LowRitmEqu.set("")
    SearchRitmEqu.set("")
    WorknoteRitmEqu.set("")
    WorknoteRitmEqu1.set("")
    ClientCountEqu.set("")
    
def ExitWindow():
    root.destroy()
    
def total():
    global TotalRitm
    for i in sheet[2:sheet.max_row]:
        TotalRitm+=1
    TotalRitmEqu.set(TotalRitm)
    TotalRitm = 0



#========State of RITM=========
def StateRitmfunc():
    global OpenCount
    global PendingCount
    global WipCount
    global OthersStateCount
    for i in range(2, (sheet.max_row)+1): 
         cell_obj = sheet.cell(row=i, column=3)
         if (cell_obj.value == "Open"):
             OpenCount+=1 
         elif(cell_obj.value == "Pending"): 
             PendingCount+=1 
         elif(cell_obj.value == "Work in Progress"): 
             WipCount+=1 
         else: 
             OthersStateCount+=1

    OpenRitmEqu.set(OpenCount)
    PendingRitmEqu.set(PendingCount)
    WipRitmEqu.set(WipCount)
    OtherStateRitmEqu.set(OthersStateCount)

    OpenCount = 0
    PendingCount=0
    WipCount=0
    OthersStateCount=0


#==All RITM Status===========
def DisplayAllRitm():
     i=0
     print("All RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i1.value != None:
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
               
     
  

#=========Priority functions============
def PriorityRitmfunc():
    global CriticalCount
    global HighCount
    global MediumCount
    global LowCount
    for i in range(2, (sheet.max_row)+1): 
         cell_obj = sheet.cell(row=i, column=6)
         if (cell_obj.value == "1 - Critical"):
             CriticalCount+=1 
         elif(cell_obj.value == "2 - High"): 
             HighCount+=1 
         elif(cell_obj.value == "3 - Medium"): 
             MediumCount+=1 
         elif(cell_obj.value == "4 - Low"): 
             LowCount+=1

    CriticalRitmEqu.set(CriticalCount)
    HighRitmEqu.set(HighCount)
    MediumRitmEqu.set(MediumCount)
    LowRitmEqu.set(LowCount)

    OthersStateCount=0
    CriticalCount=0
    HighCount=0
    MediumCount=0
    LowCount=0


#====================RITM Status Functions===================
def OpenRitmFunc():
     global OpenCount
     print("Open RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i3.value == "Open":
             OpenCount+=1
             print("{0} ---{3:.30s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     OpenRitmEqu.set(OpenCount)
     OpenCount = 0


def PendingRitmFunc():
     global PendingCount
     print("Pending RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i3.value == "Pending":
             PendingCount+=1
             print("{0} ---{3:.30s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     PendingRitmEqu.set(PendingCount)
     PendingCount=0
    

def WipRitmFunc():
     global WipCount
     print("Work In Progress RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i3.value == "Work in Progress":
             WipCount+=1
             print("{0} ---{3:.30s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     WipRitmEqu.set(WipCount)
     WipCount=0




#===========RITM Priority Functions========
     
def CriticalRitmFunc():
     global CriticalCount
     print("Critical Priority RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i6.value == "1 - Critical":
             CriticalCount+=1
             print("{0} ---{3:.30s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     CriticalRitmEqu.set(CriticalCount)
     CriticalCount=0
    
def HighRitmFunc():
     global HighCount
     print("High Priority RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i6.value == "2 - High":
             HighCount+=1
             print("{0} ---{3:.30s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     HighRitmEqu.set(HighCount)
     HighCount=0
    

def MediumRitmFunc():
     global MediumCount
     print("Medium Priority RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i6.value == "3 - Medium":
             MediumCount+=1
             print("{0} ---{3:.20s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     MediumRitmEqu.set(MediumCount)
     MediumCount=0
    

def LowRitmFunc():
     global LowCount
     print("Low Priority RITMs\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i6.value == "4 - Low":
             LowCount+=1
             print("{0} ---{3:.20s} ----{2} -------{5} --------{6} -------{8}-------{10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))

     LowRitmEqu.set(LowCount)
     LowCount=0


#===========Function to search RITM =========
def SearchRitmFunc():
     global SearchRitmRef
     SearchRitmRef = SearchRitmEqu.get()
     print("Searched RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1: 
         if i1.value == SearchRitmRef.upper():
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             SearchRitmRef=""
             
             break
             
     else:
         print("No record found")


#=========Function to check worknote=======

def WorknoteRitmFunc():
              
     global WorknoteRitmRef
     global WorknoteRitmRef1
     WorknoteRitmRef = WorknoteRitmEqu.get()
     print("Searched RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'T2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11, i12, i13, i14, i15, i16, i17, i18, i19, i20 in data1: 
         if i1.value == WorknoteRitmRef.upper():
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value, i12.value, i13.value, i14.value, i15.value, i16.value, i17.value, i18.value, i19.value, i20.value))
             WorknoteRitmRef1 = i20.value

             WorknoteRitmEqu1.set(WorknoteRitmRef1)
             WorknoteRitmRef =""
             WorknoteRitmRef1 =""
             break
             
     else:
         print("No record found")        


#==Future RITM============

def FutureScheduleRitm():
     print("Future Schedule RITM\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A2':'T2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11, i12, i13, i14, i15, i16, i17, i18, i19, i20 in data1: 
         if i9.value == None:
             continue
         elif datetime.strptime(i9.value, "%Y-%m-%d %H:%M:%S") >= datetime.strptime(dt_string, "%Y-%m-%d %H:%M:%S"):
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8:.10s} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value, i12.value, i13.value, i14.value, i15.value, i16.value, i17.value, i18.value, i19.value, i20.value))
             
    

#==Past RITM============

def PastScheduleRitm():
     print("Past Schedule RITM\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A2':'T2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11, i12, i13, i14, i15, i16, i17, i18, i19, i20 in data1: 
         if i9.value == None:
             continue
         elif datetime.strptime(i9.value, "%Y-%m-%d %H:%M:%S") <= datetime.strptime(dt_string, "%Y-%m-%d %H:%M:%S"):
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8:.10s} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value, i12.value, i13.value, i14.value, i15.value, i16.value, i17.value, i18.value, i19.value, i20.value))
             
     
#=========Clintwise RITM==========

def ABSRitmFunc():
     global z
     z=0
     print("ABS RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Ascension Care Management LLC (ABS)" or i2.value == "Advanced Benefit Solutions LLC - QL (ABS)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0

def NHPRitmFunc():
     global z
     z=0
     print("NHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Allways Health Partners (AWHP/NHP)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0 


def BRIRitmFunc():
     global z
     z=0
     print("BRI RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "BCBS of Rhode Island (BRI)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0

def BNCRitmFunc():
     global z
     z=0
     print("BNC RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "BCBS North Carolina (BNC)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0

def BCNRitmFunc():
     global z
     z=0
     print("BCN RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Blue Care Network (BCN)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0

def BRTRitmFunc():
     global z
     z=0
     print("BRT RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Bright Health Plan (BRT)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0

def CPLRitmFunc():
     global z
     z=0
     print("CPL RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Centers Plan for Healthy Living (CPL)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def CISHOCRitmFunc():
     global z
     z=0
     print("NHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "CIS Healthcare Operations Center (CISHOC)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def COARitmFunc():
     global z
     z=0
     print("COA RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Colorado Access Health (COA)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def CFHRitmFunc():
     global z
     z=0
     print("CFH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Community First Health Plan (CFH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def COVRitmFunc():
     global z
     z=0
     print("COV RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "County of Ventura (COV)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def DHPRitmFunc():
     global z
     z=0
     print("DHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Dean Health Plan (DHP)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def DHHRitmFunc():
     global z
     z=0
     print("DHH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Denver Health Medical Plan (DHH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def EMBRitmFunc():
     global z
     z=0
     print("EMB RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Emblem Health (EMB)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def FALRitmFunc():
     global z
     z=0
     print("FAL RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Fallon Community Health Plan (FAL)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def GHPRitmFunc():
     global z
     z=0
     print("GHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Gateway health plan (GHP)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def HCIRitmFunc():
     global z
     z=0
     print("HCI RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Hamaspik Choice Inc (HCI)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def HCHRitmFunc():
     global z
     z=0
     print("HCH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Healthcare Highways, Inc. - QL (HCH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def HSNRitmFunc():
     global z
     z=0
     print("HSN RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Health Services for Children with Special Needs (HSN)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def HMSRitmFunc():
     global z
     z=0
     print("HMS RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Hawaii Medical Service Assoc (HMS)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def ICHRitmFunc():
     global z
     z=0
     print("ICH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Independent Care Health Plan (ICH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def AVMRitmFunc():
     global z
     z=0
     print("AVM RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "InnovaCare (AVM)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def KMHRitmFunc():
     global z
     z=0
     print("KMH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Keystone Mercy Health Plan (KMH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def MHHRitmFunc():
     global z
     z=0
     print("MHH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Memorial Hermann Health (MHH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def LACRitmFunc():
     global z
     z=0
     print("LAC RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "LA Care Health Plan (LAC)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def PFSRitmFunc():
     global z
     z=0
     print("PFS RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "PacificSource Health Plans (PFS)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def PHPRitmFunc():
     global z
     z=0
     print("PHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Physicians Health Plan Shared Services (PHP)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def PRHRitmFunc():
     global z
     z=0
     print("PRH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Prominence Health Plan (PRH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def QCARitmFunc():
     global z
     z=0
     print("QCA RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "QualChoice of Arkansas (QCA)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def SWHRitmFunc():
     global z
     z=0
     print("SWH RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Senior Whole Health (SWH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0 


def TXCRitmFunc():
     global z
     z=0
     print("TXC RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Texas Children Health Plan (TXC)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0 




def UHGRitmFunc():
     global z
     z=0
     print("UHG RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "United HealthCare ClaimSphere (UHG)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def CMCRitmFunc():
     global z
     z=0
     print("CMC RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "UnitedHealth Group Complex Medical Conditions (CMC)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def UHARitmFunc():
     global z
     z=0
     print("UHA RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "University Health Alliance (UHA)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0




def VHPRitmFunc():
     global z
     z=0
     print("VHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Valley Health Plan (VHP)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def VPHRitmFunc():
     global z
     z=0
     print("VHP RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Virginia Premier Health (VPH)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def VIBRitmFunc():
     global z
     z=0
     print("VIB RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Vibra Health Plan (VIB)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0





def VHIRitmFunc():
     global z
     z=0
     print("VHI RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Viva Health Inc. (VHI)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0



def WMKRitmFunc():
     global z
     z=0
     print("WMK RITM Status\n") 
     print("Number -----------Short Desc-------------------------State ----- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A1':'K2000'] 
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         if i2.value == "Wellmark (WMK)":
             print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
             z+=1
     ClientCountEqu.set(z)
     z=0


def OtherRitmFunc():
     global z
     z=0
     print("Other RITM Status\n") 
     print("Number -----------Short Desc-------------------------State --------- Priority --------Opened Date --------------Prefered date ---------------Assigned")
     print("\n")
     data1 = sheet['A2':'K2000']
     for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11 in data1:
         try:
             if i2.value not in ClientName and i1.value != None:
                 print("{0} === {3:.30s} ==== {2} ======= {5} ======== {6} ======== {8} ======= {10}".format(i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value))
                 z+=1
         except:
             print("Ignore Error")
     ClientCountEqu.set(z)
     z=0



#=====Main Block=======  

if __name__ == "__main__":

    root  = tkinter.Tk()
    root.config(background = 'black')
    root.title("RITM report")
    root.geometry("1200x1200")

      

    TotalRitmEqu = StringVar()
    OpenRitmEqu = StringVar()
    PendingRitmEqu = StringVar()
    WipRitmEqu = StringVar()
    OtherStateRitmEqu = StringVar()
    CriticalRitmEqu = StringVar()
    HighRitmEqu = StringVar()
    MediumRitmEqu = StringVar()
    LowRitmEqu = StringVar()
    SearchRitmEqu = StringVar()
    WorknoteRitmEqu = StringVar()
    WorknoteRitmEqu1 = StringVar()
    ClientCountEqu = StringVar()


    #Total Count
    TotalCountLab = Label(root, fg='black', text = "Total Count", bg='ivory3', font=('Helvetica', 9, 'bold'),padx=7, pady=7,width=10,bd=2, relief=GROOVE)
    TotalCountLab.place(x= 40, y=50)

    TotalRitmLab = Label(root, text=(sheet.max_row)-1,fg='black', bg='ivory3',font=('Helvetica', 9, 'bold'), padx=7, pady=7,width=10, bd=2, relief=GROOVE)
    TotalRitmLab.place(x= 40, y=100)



    #All RITM, Clear All, Exit
    TotalRitmBut = Button(root, text="All RITM", fg='white', bg='black',font=('Helvetica', 9, 'bold'), padx=7, pady=7,width=10, command=DisplayAllRitm)
    TotalRitmBut.place(x= 40, y=150)
    
    ClearAllRitmBut = Button(root, text="ClearAll", fg='white', bg='black',font=('Helvetica', 9, 'bold'), padx=7, pady=7,width=10, command=clear)
    ClearAllRitmBut.place(x= 40, y=200)

    ExitBut = Button(root, text="Exit", fg='white', bg='black',font=('Helvetica', 9, 'bold'), padx=7, pady=7,width=10, command=ExitWindow)
    ExitBut.place(x= 40, y=250)

    # Search RITM
    CheckSingleRitmBut = Button(root, text="Search RITM", fg='white', bg='lightcyan4', padx=7, pady=5,width=14, command=SearchRitmFunc)
    CheckSingleRitmBut.place(x= 180, y=250)
    CheckSingleRitmEntry = Entry(root, textvariable=SearchRitmEqu)
    CheckSingleRitmEntry.place(x= 180, y=280)

     # Check RITM Comments and Work notes
    WorkNoteRitmBut = Button(root, text="Check Work Note", fg='white', bg='lightcyan4', padx=7, pady=5,width=14, command=WorknoteRitmFunc)
    WorkNoteRitmBut.place(x= 300, y=250)
    WorkNoteRitmEntry = Entry(root, textvariable=WorknoteRitmEqu)
    WorkNoteRitmEntry.place(x= 300, y=280)
    worknoteRitmLabe2 = Label(root,font=('Helvetica', 9, 'bold'), text="Comments and Worknotes", padx=10).place(x=590, y=10)
    worknoteRitmLabel = Label(root, textvariable=WorknoteRitmEqu1).place(x=590, y=40)

     # Futre and past Schedule of RITM
    FutureRitmBut = Button(root, text="Future Schedule", fg='white', bg='lightcyan4',font=('Helvetica', 9, 'bold'), padx=7, pady=5,width=14, command=FutureScheduleRitm)
    FutureRitmBut.place(x= 440, y=450)
    PastRitmBut = Button(root, text="Past Schedule", fg='white', bg='lightcyan4',font=('Helvetica', 9, 'bold'), padx=7, pady=5,width=14, command=PastScheduleRitm)
    PastRitmBut.place(x= 440, y=500)
    

    
    #State of RITM

    StatusRitmBut = Button(root, text="RITM Status Count", fg='white', bg='darkgoldenrod', padx=15, pady=3,width=10, command=StateRitmfunc)
    StatusRitmBut.place(x= 270, y=20)
    
    OpenRitmBut = Button(root, text="Open RITM", fg='white', bg='lightcyan4', padx=7, pady=7,width=10, command=OpenRitmFunc)
    OpenRitmBut.place(x= 180, y=50)
    OpenRitmEntry = Entry(root, textvariable=OpenRitmEqu).place(x=180, y=80)

    PendingRitmBut = Button(root, text="Pending RITM", fg='white', bg='lightcyan4', padx=7, pady=7,width=10, command=PendingRitmFunc)
    PendingRitmBut.place(x= 280, y=50)
    PendingRitmEntry = Entry(root, textvariable=PendingRitmEqu).place(x=280, y=80)

    WipRitmBut = Button(root, text="WorkInProgress", fg='white', bg='lightcyan4', padx=7, pady=7,width=10, command=WipRitmFunc)
    WipRitmBut.place(x=380, y=50)
    WipRitmButEntry = Entry(root, textvariable=WipRitmEqu).place(x=380, y=80)

    
    #Priority of RITM

    PriorityRitmBut = Button(root, text="RITM Priority Count", fg='white', bg='darkgoldenrod', padx=15, pady=3,width=11, command=PriorityRitmfunc)
    PriorityRitmBut.place(x= 280, y=130)
    
    CriticalRitmBut = Button(root, text="1 - Critical", fg='white', bg='lightcyan4', padx=7, pady=7,width=10, command=CriticalRitmFunc)
    CriticalRitmBut.place(x= 180, y=160)
    CriticalRitmEntry = Entry(root, textvariable=CriticalRitmEqu).place(x=180, y=190)

    HighRitmBut = Button(root, text="2 - High", fg='white', bg='lightcyan4', padx=7, pady=7,width=10, command=HighRitmFunc)
    HighRitmBut.place(x= 270, y=160)
    HighRitmEntry = Entry(root, textvariable=HighRitmEqu).place(x=270, y=190)

    MediumRitmBut = Button(root, text="3 - Medium", fg='white', bg='lightcyan4', padx=7, pady=7,width=10, command=MediumRitmFunc)
    MediumRitmBut.place(x=360, y=160)
    MediumRitmButEntry = Entry(root, textvariable=MediumRitmEqu).place(x=360, y=190)

    LowRitmBut = Button(root, text="4 - Low", fg='white', bg='lightcyan4', padx=7, pady=7,width=7, command=LowRitmFunc)
    LowRitmBut.place(x=450, y=160)
    LowRitmEntry = Entry(root,width=11, textvariable=LowRitmEqu).place(x=450, y=190)
    
    #======Print date and time======
    DateAndTimeLabel = Label(root,font=('Helvetica', 9, 'bold'), text=("date and time = ",datetime.strptime(dt_string, "%Y-%m-%d %H:%M:%S")))
    DateAndTimeLabel.place(x=900, y=10)

    
    #====Clientwise RITM========

    AllClientButton = Button(root, text="All Clients", padx=60, width=9)
    AllClientButton.place(x=80, y=370)

    ClientCount = Entry(root, width=9,textvariable=ClientCountEqu)
    ClientCount.place(x=300, y=370)

    AbsButton = Button(root, text="ABS", padx=8, width=8, command=ABSRitmFunc)
    AbsButton.place(x=50, y=400)

    NhpButton = Button(root, text="AWHP/NHP", padx=8, width=8, command=NHPRitmFunc)
    NhpButton.place(x=50, y=430)

    BriButton = Button(root, text="BRI", padx=8, width=8, command=BRIRitmFunc)
    BriButton.place(x=50, y=460)

    BcnButton = Button(root, text="BCN", padx=8, width=8, command=BCNRitmFunc)
    BcnButton.place(x=50, y=490)

    BrtButton = Button(root, text="BRT", padx=8, width=8, command=BRTRitmFunc)
    BrtButton.place(x=50, y=520)

    BncButton = Button(root, text="BNC", padx=8, width=8, command=BNCRitmFunc)
    BncButton.place(x=50, y=550)

    CoaButton = Button(root, text="COA", padx=8, width=8, command=COARitmFunc)
    CoaButton.place(x=50, y=580)

    CfhButton = Button(root, text="CFH", padx=8, width=8, command=CFHRitmFunc)
    CfhButton.place(x=50, y=610)

    CplButton = Button(root, text="CPL", padx=8, width=8, command=CPLRitmFunc)
    CplButton.place(x=50, y=640)

    CishocButton = Button(root, text="CISHOC", padx=8, width=8, command=CISHOCRitmFunc)
    CishocButton.place(x=50, y=670)
        
    CovButton = Button(root, text="COV", padx=8, width=8, command=COVRitmFunc)
    CovButton.place(x=130, y=400)

    DhpButton = Button(root, text="DHP", padx=8, width=8, command=DHPRitmFunc)
    DhpButton.place(x=130, y=430)

    DhhButton = Button(root, text="DHH", padx=8, width=8, command=DHHRitmFunc)
    DhhButton.place(x=130, y=460)

    EmbButton = Button(root, text="EMB", padx=8, width=8, command=EMBRitmFunc)
    EmbButton.place(x=130, y=490)

    FalButton = Button(root, text="FAL", padx=8, width=8, command=FALRitmFunc)
    FalButton.place(x=130, y=520)

    GhpButton = Button(root, text="GHP", padx=8, width=8, command=GHPRitmFunc)
    GhpButton.place(x=130, y=550)

    HciButton = Button(root, text="HCI", padx=8, width=8, command=HCIRitmFunc)
    HciButton.place(x=130, y=580)

    HmsButton = Button(root, text="HMS", padx=8, width=8, command=HMSRitmFunc)
    HmsButton.place(x=130, y=610)

    HsnButton = Button(root, text="HSN", padx=8, width=8, command=HSNRitmFunc)
    HsnButton.place(x=130, y=640)

    HchButton = Button(root, text="HCH", padx=8, width=8, command=HCHRitmFunc)
    HchButton.place(x=130, y=670)

    IchButton = Button(root, text="ICH", padx=8, width=8, command=ICHRitmFunc)
    IchButton.place(x=210, y=400)

    AvmButton = Button(root, text="AVM", padx=8, width=8, command=AVMRitmFunc)
    AvmButton.place(x=210, y=430)

    KmhButton = Button(root, text="KMH", padx=8, width=8, command=KMHRitmFunc)
    KmhButton.place(x=210, y=460)

    LacButton = Button(root, text="LAC", padx=8, width=8, command=LACRitmFunc)
    LacButton.place(x=210, y=490)

    MhhButton = Button(root, text="MHH", padx=8, width=8, command=MHHRitmFunc)
    MhhButton.place(x=210, y=520)

    PfsButton = Button(root, text="PFS", padx=8, width=8, command=PFSRitmFunc)
    PfsButton.place(x=210, y=550)

    PhpButton = Button(root, text="PHP", padx=8, width=8, command=PHPRitmFunc)
    PhpButton.place(x=210, y=580)

    PrhButton = Button(root, text="PRH", padx=8, width=8, command=PRHRitmFunc)
    PrhButton.place(x=210, y=610)

    QcaButton = Button(root, text="QCA", padx=8, width=8, command=QCARitmFunc)
    QcaButton.place(x=210, y=640)

    SwhButton = Button(root, text="SWH", padx=8, width=8, command=SWHRitmFunc)
    SwhButton.place(x=210, y=670)

    TxcButton = Button(root, text="TXC", padx=8, width=8, command=TXCRitmFunc)
    TxcButton.place(x=290, y=400)

    UhgButton = Button(root, text="UHG", padx=8, width=8, command=UHGRitmFunc)
    UhgButton.place(x=290, y=430)

    CmcButton = Button(root, text="CMC", padx=8, width=8, command=CMCRitmFunc)
    CmcButton.place(x=290, y=460)

    UhaButton = Button(root, text="UHA", padx=8, width=8, command=UHARitmFunc)
    UhaButton.place(x=290, y=490)

    VphButton = Button(root, text="VHP", padx=8, width=8, command=VPHRitmFunc)
    VphButton.place(x=290, y=520)

    VibButton = Button(root, text="VIB", padx=8, width=8, command=VIBRitmFunc)
    VibButton.place(x=290, y=550)

    VphButton = Button(root, text="VPH", padx=8, width=8, command=VPHRitmFunc)
    VphButton.place(x=290, y=580)

    VhiButton = Button(root, text="VHI", padx=8, width=8, command=VHIRitmFunc)
    VhiButton.place(x=290, y=610)

    WmkButton = Button(root, text="WMK", padx=8, width=8, command=WMKRitmFunc)
    WmkButton.place(x=290, y=640)

    OtherClientButton = Button(root, text="Others", padx=8, width=8, command=OtherRitmFunc)
    OtherClientButton.place(x=290, y=670)

    

    root.mainloop()