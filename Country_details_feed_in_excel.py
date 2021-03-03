from openpyxl import *
from tkinter import *

wrbk = load_workbook('excel.xlsx')

sheet = wrbk.active

def excel1():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30

    sheet.cell(row=1, column=1).value = "Country"
    sheet.cell(row=1, column=2).value = "Capital"
    sheet.cell(row=1, column=3).value = "Religion"
    sheet.cell(row=1, column=4).value = "Population"
    sheet.cell(row=1, column=5).value = "Area"
    sheet.cell(row=1, column=6).value = "GDP"
    sheet.cell(row=1, column=7).value = "Sport"
    sheet.cell(row=1, column=8).value = "Climate"
    sheet.cell(row=1, column=9).value = "Main River"



def clear():
    Country_field.delete(0, END)
    Capital_field.delete(0, END)
    Religion_field.delete(0, END)
    Population_field.delete(0, END)
    Area_field.delete(0, END)
    GDP_field.delete(0, END)
    Sport_field.delete(0, END)
    Climate_field.delete(0, END)
    Main_River_field.delete(0, END)




def insert():
    if Country_field.get() =="" and Capital_field.get() == "" and Religion_field.get() == "" and Population_field.get() == ""\
       and Area_field.get() == "" and GDP_field.get() == "" and Sport_field.get() == "" \
       and Climate_field.get() == "" and Main_River_field.get() == "":
        print("Empty Input")

    else:
        current_row = sheet.max_row
        sheet.cell(row=current_row+1, column=1).value = Country_field.get()
        sheet.cell(row=current_row+1, column=2).value = Capital_field.get()
        sheet.cell(row=current_row+1, column=3).value = Religion_field.get()
        sheet.cell(row=current_row+1, column=4).value = Population_field.get()
        sheet.cell(row=current_row+1, column=5).value = Area_field.get()
        sheet.cell(row=current_row+1, column=6).value = GDP_field.get()
        sheet.cell(row=current_row+1, column=7).value = Sport_field.get()
        sheet.cell(row=current_row+1, column=8).value = Climate_field.get()
        sheet.cell(row=current_row+1, column=9).value = Main_River_field.get()
        
        
        wrbk.save('excel.xlsx')


        clear()



if __name__ == "__main__":
    base = Tk()

    base.configure(background="light yellow")

    base.title("Country's details form")

    base.geometry("500x300")


    excel1()

    Country = Label(base, text="Country", bg="blue", fg="white")
    Capital = Label(base, text="Capital", bg="blue", fg="white")
    Religion = Label(base, text="Religion", bg="blue", fg="white")
    Population = Label(base, text="Population", bg="blue", fg="white")
    Area = Label(base, text="Area", bg="blue", fg="white")
    GDP = Label(base, text="GDP", bg="blue", fg="white")
    Sport = Label(base, text="Sport", bg="blue", fg="white")
    Climate = Label(base, text="Climate", bg="blue", fg="white")
    Main_River = Label(base, text="Main_River", bg="blue", fg="white")

    Country.grid(row=1, column=0)
    Capital.grid(row=2, column=0)
    Religion.grid(row=3, column=0)
    Population.grid(row=4, column=0)
    Area.grid(row=5, column=0)
    GDP.grid(row=6, column=0)
    Sport.grid(row=7, column=0)
    Climate.grid(row=8, column=0)
    Main_River.grid(row=9, column=0)



    Country_field =Entry(base)
    Capital_field = Entry(base)
    Religion_field = Entry(base)
    Population_field = Entry(base)
    Area_field = Entry(base)
    GDP_field = Entry(base)
    Sport_field = Entry(base)
    Climate_field = Entry(base)
    Main_River_field = Entry(base)


    Country_field.grid(row=1, column=1, ipadx="100")
    Capital_field.grid(row=2, column=1, ipadx="100")
    Religion_field.grid(row=3, column=1, ipadx="100")
    Population_field.grid(row=4, column=1, ipadx="100")
    Area_field.grid(row=5, column=1, ipadx="100")
    GDP_field.grid(row=6, column=1, ipadx="100")
    Sport_field.grid(row=7, column=1, ipadx="100")
    Climate_field.grid(row=8, column=1, ipadx="100")
    Main_River_field.grid(row=9, column=1, ipadx="100")

    excel1()

    submit = Button(base, text="Submit", fg="Black", bg="Red", command=insert)
    submit.grid(row=11, column=1)

    base.mainloop()

    
    
